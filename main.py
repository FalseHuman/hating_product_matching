import re
import openpyxl
import logging
import json
import time
from fuzzywuzzy import fuzz
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from utils import clean_price_data, create_driver


def translate_color_in_excel(file_path):
    """
    Читает данные из Excel-файла, заменяет английские названия цветов на русские.

    Args:
        file_path (str): Путь к Excel-файлу.

    Returns:
        bool: True, если замена произведена успешно, False - если нет
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        color_translations = {
            'black': 'черный',
            'white': 'белый',
            'gray': 'серый',
            'silver': 'серебристый',
            'gold': 'золотой',
            'blue': 'синий',
            'green': 'зеленый',
            'red': 'красный',
            'purple': 'фиолетовый',
            'pink': 'розовый',
            'midnight': 'темная ночь',
            'starlight': 'сияющая звезда',
            'space black': 'космический черный',
            'deep purple': 'темно-фиолетовый',
            'alpine green': 'альпийский зеленый',
            'sierra blue': 'небесно-голубой',
            'yellow': 'желтый',
            'graphite': 'графитовый',
            'lavender': 'лавандовый',
            'mint': 'мятный',
            'lilac': 'сиреневый',
            'lemon': 'лимонный',
            'obsidian': 'обсидиан',
            'hazel': 'ореховый',
            'porcelain': 'фарфоровый',
            'peony': 'пионовый',
            'rosequartz': 'розовый кварц',
            'olive': 'оливковый',
            'natural': 'натуральный',
            'tan': 'загорелый',
            'platinum': 'платиновый',
            'sky': 'небесный',
            'denim': 'деним',
            'cloud': 'облачный',
            'blush': 'румяна',
            'ink': 'чернила',
            'winter blue': 'зимний синий',
            'sea': 'морской',
            'charcoal': 'угольный',
            'lemongrass': 'лимонник',
            'stormy black': 'штормовой черный',
            'cloudy white': 'облачный белый',
            'ocean blue': 'океанский синий',
            'lime green': 'лаймовый зеленый',
            'pure platinum': 'чистая платина',
            'midnight sky': 'полуночное небо',
            'rose gold': 'розовое золото',
            'plum': 'сливовый',
            'teal': 'бирюзовый',
            'ultramarine': 'ультрамариновый',
            'desert': 'пустынный'
        }

        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                   text = str(cell.value)
                   # Ищем все вхождения английских названий цветов в строке
                   for english_color, russian_color in color_translations.items():
                       pattern = r'\b' + re.escape(english_color) + r'\b'
                       text = re.sub(pattern, russian_color,
                                     text, flags=re.IGNORECASE)

                   cell.value = text
        file_path = file_path.split('.')
        file_path[0] = '_' + file_path[0] + '_'
        new_file_path = '.'.join(file_path)
        workbook.save(new_file_path)
        return True, new_file_path
    except FileNotFoundError:
        print(f"Ошибка: Файл не найден: {file_path}")
        return False
    except Exception as e:
        print(f"Произошла ошибка при чтении или записи файла: {e}")
        return False


def save_file(data, file):
    try:
        with open(file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
    except IOError as e:
        logging.error(f"Произошла ошибка ввода-вывода: {e}")
    except TypeError as e:
        logging.error(f"Ошибка типа данных: {e}")
    except Exception as e:
        logging.error(f"Произошла непредвиденная ошибка: {e}")


def open_json_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            data = json.load(file)
            return data
    except FileNotFoundError:
        logging.error(f"Файл {file_path} не найден.")
    except json.JSONDecodeError:
        logging.error(f"Ошибка декодирования JSON в файле {file_path}.")
    except Exception as e:
        logging.error(f"Произошла ошибка: {e}")


def read_excel_file(file_path):
    """
    Читает данные из Excel-файла и возвращает их.

    Args:
      file_path (str): Путь к Excel-файлу.

    Returns:
      list or None: Список списков (строки) или None, если произошла ошибка.
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active  # Получаем активный лист

        data = []
        # Получаем строки как списки
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

        return data
    except FileNotFoundError:
        print(f"Ошибка: Файл не найден: {file_path}")
        return None
    except Exception as e:
        print(f"Произошла ошибка при чтении файла: {e}")
        return None


def extract_color_from_string(text):
    """
    Извлекает цвет из строки, если он там есть.
    Args:
        text (str): Строка, в которой нужно найти цвет.
    Returns:
        str: Цвет (в нижнем регистре) или None, если цвет не найден.
    """
    color_matches = re.findall(r'\b(Black|White|Gray|Silver|Gold|Blue|Green|Red|Purple|Pink|Midnight|Starlight|Space Black|Deep Purple|Alpine Green|Sierra Blue|Yellow|Graphite|Lavender|Mint|Lilac|Lemon|Obsidian|Hazel|Porcelain|Peony|Rose Quartz|Olive|Natural|Tan|Platinum|Sky|Denim|Cloud|Blush|Ink|Winter Blue|Sea|Charcoal|Lemongrass|Stormy Black|Cloudy White|Ocean Blue|Lime Green|Pure Platinum|Midnight Sky|Rose Gold|Plum|Teal|Ultramarine|Desert|черный|белый|серый|серебристый|золотой|синий|зеленый|красный|фиолетовый|розовый|темная ночь|сияющая звезда|космический черный|темно-фиолетовый|альпийский зеленый|небесно-голубой|желтый|графитовый|лавандовый|мятный|лимонный|обсидиан|ореховый|фарфоровый|пионовый|розовый кварц|оливковый|натуральный|загорелый|платиновый|небесный|деним|облачный|румяна|чернила|зимний синий|морской|угольный|лимонник|штормовой черный|облачный белый|океанский синий|лаймовый зеленый|чистая платина|полуночное небо|розовое золото|сливовый|бирюзовый|ультрамариновый|пустынный)\b', text, re.IGNORECASE)
    if color_matches:
        return color_matches[0].lower()
    return None


def extract_model_number(model_string):
    """
    Извлекает номер модели из строки.

    Args:
        model_string: Строка с названием модели.

    Returns:
        str: Номер модели или None, если не найден.
    """
    match = re.search(r'(\d{1,3}(?:\s*[A-Za-z+]*\s*\d{0,3})*)',
                      model_string)  # Ищем 1 или 3 цифры, возможно с буквами между ними
    if match:
        return match.group(1)
    return None


def match_products(device_data_list, database_products):
    """
    Сопоставляет товары из списка словарей с товарами из магазина.

    Args:
        device_data_list (list): Список словарей с данными об устройствах
          (в формате  {'model': 'Honor 200 Lite 8/256 Green -', 'price': 17800, 'customer': '112пав'}).
        database_products (list): Список словарей с товарами из магазина.

    Returns:
        tuple: Словарь с сопоставленными и не сопоставленными товарами.
    """
    matched_products = {}
    unmatched_products = []

    for device_data in device_data_list:
        best_match = None
        best_score = 0

        for db_product in database_products:
            score = 0
            try:
                if 'yandex' in device_data:
                    if device_data['yandex'] is not None:
                        db_product_ = db_product[3] + ' ' + \
                            db_product[8] + ' ' + db_product[10]
                        score = fuzz.partial_ratio(
                            device_data['yandex'].lower(), db_product_.lower())
                        if score > best_score and score > 80:  # Устанавливаем порог сходства
                            best_score = score
                            best_match = db_product
            except:
                pass

        if best_match:
            data = {**device_data, "matched": True, "db_product": best_match}
            if matched_products.get(device_data['model'], None) is None:
                matched_products[device_data['model']] = [data]
            else:
                if data is not matched_products[device_data['model']]:
                    matched_products[device_data['model']].append(data)
        else:
            unmatched_products.append({**device_data, "matched": False})

    return matched_products, unmatched_products


def print_results(matched, unmatched, set_customer_list, matched_file="matched_products.xlsx", output_file="unmatched_products.xlsx"):
    """
    Выводит результаты сопоставления.

    Args:
        matched (dict): Словарь с сопоставленными товарами.
        unmatched (list): Список  несопоставленных словарей.
    """
    print("Сопоставленные товары:", len(matched))

    if matched:
      try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # print(set_customer_list)
        recommended_price_array_head = [
            "Рекомендуемая цена"] * len(set_customer_list)
        recommended_prices_head = []
        max_customer = max([len(product)for product in matched.values()])
        # print(max_customer)
        for i in range(0, max_customer):
            supplier = f"Поставщик {i + 1}"
            price = f"Цена поставщика {i + 1}"
            recommended_prices_head.append(supplier)
            recommended_prices_head.append(price)
        headers = ['Наименование', 'Внешний код', 'Производитель', 'Модель',
                   'Оперативная память (Gb)', 'Кол-во симкарт', 'Тип аппарата',	'Процессор', 'Цвет', 'Код производителя', 'Встроенная память'] + recommended_price_array_head
        for rec_p in recommended_prices_head:
            headers.append(rec_p)
        sheet.append(headers)  # добавляем заголовок
        for product in matched.values():
            customer_prices = []
            recommended_prices = [""] * len(set_customer_list)
            if len(product) > 1:
                min_price = []
                for customer in set_customer_list:
                    try:
                        index = set_customer_list.index(customer)
                        # print(product[index])
                        min_price.append(product[index]['price'])
                        if min(min_price) not in recommended_prices:
                            recommended_prices[index] = min(
                                min_price) + (min(min_price)*0.1)
                        customer_prices.append(product[index]['customer'])
                        customer_prices.append(product[index]['price'])
                    except (ValueError, IndexError):
                        customer_prices.append(' ')
                        customer_prices.append(' ')

            else:
                product_ = product[0]
                # print(product_)
                index = set_customer_list.index(product_['customer'])
                recommended_prices = [""] * len(set_customer_list)
                recommended_prices[index] = product_[
                    'price'] + (product_['price']*0.1)
                customer_prices.append(product_['customer'])
                customer_prices.append(product_['price'])

            product = product[0].get('db_product')

            sheet.append([*list(product), *recommended_prices,
                         *customer_prices])  # добавляем данные
        workbook.save(matched_file)
        print(f"\nСопоставленные товары сохранены в '{matched_file}'")
      except Exception as e:
        print(f"Ошибка при записи в Excel с сопоставленными товарами: {e}")

    print("\nНе сопоставленные товары:",  len(unmatched))
    if unmatched:
      try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Модель", "Цена", "Продавец"])  # добавляем заголовок
        for product in unmatched:
            sheet.append([product['model'], product['price'],
                         product['customer']])  # добавляем данные
        workbook.save(output_file)
        print(f"\nНесопоставленные товары сохранены в '{output_file}'")
      except Exception as e:
        print(f"Ошибка при записи в Excel: {e}")


if __name__ == "__main__":
    file_path = input('Введите файл с товарами, например Прайсы с телеграма 28.01.xlsx\n')
    success, new_file_path = translate_color_in_excel(file_path)
    if success:
        file_path = new_file_path
        print(f"Цвета в файле '{file_path}' успешно заменены на русские.")
    else:
        print(f"Не удалось произвести замену цветов в файле '{file_path}'.")
    cleaned_data = clean_price_data(file_path)
    shop_items_file_path = input('Введите файл с товарами, например Товары магазина.xlsx\n')
    shop_items_data = read_excel_file(shop_items_file_path)
    customer_list = [list_['customer']
                     for list_ in cleaned_data if list_.get('customer', None)]
    set_customer_list = list(set(customer_list))  # Уникальные продавцы
    # Если нужно проверить данные на Маркете нужно расскомментировать код ниже
    if cleaned_data:
        print("Очищенные данные:")
        count = 0
        for item in cleaned_data:
            count += 1
            print(f'{count}/{len(cleaned_data)}', item['model'])
            item['yandex'] = 'yandex'
            time.sleep(5)
            try:
                driver = create_driver()
                driver.get('https://market.yandex.ru/search?text=' + item['model'])
                div1 = driver.find_element(
                By.XPATH, "//div[@data-auto-themename='listDetailed']")
                span = div1.find_element(By.XPATH, ".//span[@data-auto='snippet-title']")
                item['yandex'] = span.text
            except NoSuchElementException:
                item['yandex'] = None
            finally:
                driver.close()
                driver.quit()
        save_file(file='res.json', data=cleaned_data)
    else:
        print("Не удалось прочитать файл или очистить данные.")
    if shop_items_data:
        cleaned_data = open_json_file(file_path='res.json')
        matched_products, unmatched_products = match_products(
            cleaned_data, shop_items_data)
        print_results(matched_products, unmatched_products, set_customer_list)
    else:
        print("Не удалось прочитать файл.")
